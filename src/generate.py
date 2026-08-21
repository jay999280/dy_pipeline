# -*- coding: utf-8 -*-
"""⑤ 生成：按赛道批量生成脚本，输出「对标链接｜发布文案｜画面｜文案」Excel。"""
import concurrent.futures as cf
import difflib
import json
import logging
import os
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook

from analyze import call_deepseek, get_api_key
from common import DATA, load_card, read_json, run_dir, setup_log, write_json

try:
    from rapidfuzz import fuzz as _rfuzz
except ImportError:
    _rfuzz = None

log = logging.getLogger(__name__)


# ---------- few-shot：解析人工脚本 xlsx ----------
def parse_fewshot_xlsx(path: str) -> list:
    """解析已有脚本 Excel：发布文案 + 逐镜头(画面/文案)。

    真实格式（人工脚本示例）：R0 对标链接；R1 发布文案+正文；
    R2 画面/文案表头（文案列=ti，画面数据在 ti 左侧一列）；
    数据行画面为空的行是上一镜头的文案续行。
    """
    if not path or not Path(path).exists():
        log.warning("few-shot 脚本文件不存在: %s", path)
        return []
    examples = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            caption, hdr_idx, ti = "", None, None
            for ri, row in enumerate(rows):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if not caption and "发布文案" in cells:
                    ci = cells.index("发布文案")
                    caption = next((c for c in cells[ci + 1:] if c), "")
                if hdr_idx is None and "画面" in cells and "文案" in cells:
                    hdr_idx, ti = ri, cells.index("文案")
                    break
            shots = []
            if hdr_idx is not None:
                for row in rows[hdr_idx + 1:]:
                    if ti >= len(row) or row[ti] in (None, ""):
                        continue
                    text = str(row[ti]).strip()
                    pic = ""
                    for j in range(ti - 1, -1, -1):
                        if j < len(row) and row[j] not in (None, ""):
                            pic = str(row[j]).strip()
                            break
                    if pic:
                        shots.append({"画面": pic, "文案": text})
                    elif shots:  # 续行，并入上一镜头
                        shots[-1]["文案"] += text
            if shots:
                examples.append({"发布文案": caption or ws.title, "镜头": shots[:16]})
        wb.close()
    except Exception as e:
        log.warning("few-shot 解析失败（跳过）: %s", e)
    log.info("few-shot 解析到 %d 组人工脚本", len(examples))
    return examples


# ---------- 生成 ----------
def parse_duration(card: dict) -> tuple:
    """解析需求卡的脚本时长要求（秒），默认 30-60。"""
    dur = str(card.get("生成设置", {}).get("脚本时长秒", "30-60"))
    nums = [int(x) for x in re.findall(r"\d+", dur)]
    lo, hi = nums[0] if nums else 30, nums[1] if len(nums) > 1 else 60
    if lo >= hi:
        hi = lo + 30
    return lo, hi


def _script_chars(s: dict) -> int:
    return sum(len(sh.get("文案", "")) for sh in s.get("镜头", []))


def _hard_trim(s: dict, cap: int) -> bool:
    """确定性裁剪超长脚本：保留首镜(钩子)与尾镜(CTA)，从中间镜头逐句删尾部。"""
    shots = s.get("镜头", [])
    if len(shots) <= 2:
        return False
    total = _script_chars(s)
    guard = 0
    while total > cap and guard < 500:
        guard += 1
        mids = [sh for sh in shots[1:-1] if len(sh.get("文案", "")) > 10]
        if not mids:
            return _script_chars(s) <= cap
        target_sh = max(mids, key=lambda sh: len(sh.get("文案", "")))
        text = target_sh.get("文案", "")
        # 过滤末尾空串（lookbehind 分割会产生空元素），保证 pop 删的是真句子
        sentences = [p for p in re.split(r"(?<=[。！？!?])", text) if p]
        if len(sentences) <= 1:
            # 无句界可切，硬删一半
            target_sh["文案"] = text[: len(text) // 2].strip()
        else:
            sentences.pop()
            target_sh["文案"] = "".join(sentences).strip()
        total = _script_chars(s)
    return _script_chars(s) <= cap


# 平铺开场检测：自我介绍式开头 或 无任何强钩子信号 → 视为弱钩子
# 注意：客户专属平铺开场词已迁入需求卡"平铺开场黑名单"字段，
# 这里只保留通用兜底；fix_hooks 会把需求卡里的词并入。
_FLAT_OPENERS = ("大家好", "大家好我是", "今天来给大家", "今天给大家",
                 "别担心", "嗨", "hello", "哈喽", "hi", "第一步", "首先")
_STRONG_SIGNALS = ("？", "！", "!", "?", "别", "坑", "骗", "后悔", "秘密", "真相",
                   "担心", "千万别", "一定要", "省钱", "上当", "吃亏", "竟然",
                   "知道吗", "为什么", "怎么", "你家的", "你是不是")

# 通用 CTA 兜底（客户可在需求卡 生成设置.CTA选项库 覆盖）
_DEFAULT_CTA = (
    "私信我，免费上门测量，先量再定方案",
    "评论区扣1，我发你本地实拍案例",
    "来展厅实地看，材质工艺随便验",
    "点关注，下期讲X",
    "收藏备用，装修时用得上",
)


def _flat_openers(card: dict) -> tuple:
    """通用平铺开场黑名单 + 需求卡自定义黑名单。"""
    extra = tuple(str(x).strip() for x in (card.get("平铺开场黑名单") or []) if str(x).strip())
    return extra + _FLAT_OPENERS


def _cta_lib(card: dict) -> list:
    """CTA 选项库：需求卡优先，缺省用通用兜底。"""
    lib = card.get("生成设置", {}).get("CTA选项库") or []
    return [str(x) for x in lib if str(x).strip()] or list(_DEFAULT_CTA)


def _is_weak_hook(text: str, flat_openers=_FLAT_OPENERS) -> bool:
    if any(text.startswith(p) for p in flat_openers):
        return True
    return not any(w in text for w in _STRONG_SIGNALS)


def fix_hooks(scripts_by_track: dict, analysis_map: dict, card: dict, api_key: str):
    """钩子兜底：首镜头是平铺开场时，按参考视频的钩子类型重写。"""
    openers = _flat_openers(card)
    for tname, scripts in scripts_by_track.items():
        for i, s in enumerate(scripts, 1):
            shots = s.get("镜头", [])
            if not shots:
                continue
            first = str(shots[0].get("文案", "")).strip()
            if not _is_weak_hook(first, openers):
                continue
            ref = analysis_map.get(str(s.get("参考视频", "")), {})
            hook = ref.get("钩子设计") or {}
            lo, hi = parse_duration(card)
            prompt = f"""你是抖音短视频钩子专家。下面这条脚本的首个镜头文案是平铺开场（自我介绍式），没有钩子。
请只重写【首镜头文案】为强钩子，要求：
1. 类型复刻参考视频的钩子：{hook.get('类型', '反常识/痛点/悬念')}
2. 参考其开场手法（原话：{hook.get('前3秒原话', '')}），但不得抄袭原句
3. 结合客户业务（{card.get('业务简介', '').strip()}）和本脚本角度（{s.get('改编角度', '')}）
4. 25~50 字，口语化、有现场感，禁止以"大家好/我是/今天给大家"开头
只输出 JSON：{{"文案":"重写后的首镜头文案"}}

原首镜头文案：{first}"""
            try:
                result = call_deepseek([{"role": "user", "content": prompt}], api_key, temperature=0.8)
                new_first = str(result.get("文案", "")).strip()
                if 15 <= len(new_first) <= 60:
                    shots[0]["文案"] = new_first
                    log.info("[%s]#%d 钩子已重写: %s", tname, i, new_first[:40])
                else:
                    log.warning("[%s]#%d 钩子重写长度异常(%d字)，保留原样",
                                tname, i, len(new_first))
            except Exception as e:
                log.warning("[%s]#%d 钩子重写失败: %s", tname, i, e)


def fix_lengths(scripts_by_track: dict, card: dict, api_key: str):
    """时长兜底：过短脚本 LLM 扩写；超长脚本确定性硬裁剪（保留钩子与 CTA，不调 LLM）。"""
    lo, hi = parse_duration(card)
    target = (lo + hi) * 2  # 目标中间字数（约 180）
    for tname, scripts in scripts_by_track.items():
        for i, s in enumerate(scripts, 1):
            n = _script_chars(s)
            if lo * 4 <= n <= hi * 4:
                continue
            if n > hi * 4:
                # 超长：确定性硬裁剪，不依赖 LLM
                _hard_trim(s, hi * 4 - 10)
                n3 = _script_chars(s)
                log.info("[%s]#%d 超长(%d字≈%.0f秒)，硬裁剪后 %d 字 ≈ %.0f 秒",
                         tname, i, n, n / 4, n3, n3 / 4)
                continue
            log.info("[%s]#%d 过短(%d字≈%.0f秒)，LLM 扩写中...", tname, i, n, n / 4)
            try:
                prompt = f"""你是短视频编导。下面这条脚本口播文案 {n} 字，不足 {lo*4} 字（约 {lo} 秒）。
请把它扩写到 {target} 字左右（上限 {hi*4} 字）：保持镜头数量和画面不变或微调，每个镜头的文案加长（补充细节、例子、口语衔接），不改变前3秒钩子和结尾CTA的结构，不照抄原文。
只输出一个 JSON 对象，结构与输入完全一致：{{"发布文案":"...","参考视频":"...","镜头":[{{"画面":"...","文案":"..."}}]}}

输入脚本：{json.dumps(s, ensure_ascii=False)}"""
                fixed = call_deepseek([{"role": "user", "content": prompt}], api_key, temperature=0.6)
                n2 = _script_chars(fixed)
                if n2 > hi * 4:
                    # 扩写过头 → 硬裁剪兜回上限内（保留钩子与 CTA）
                    _hard_trim(fixed, hi * 4 - 10)
                    n2 = _script_chars(fixed)
                if lo * 4 <= n2 <= hi * 4 and fixed.get("发布文案") and fixed.get("镜头"):
                    scripts[i - 1] = fixed
                    log.info("[%s]#%d 扩写后 %d 字 ≈ %.0f 秒", tname, i, n2, n2 / 4)
                else:
                    log.warning("[%s]#%d 扩写未达标(%d字)，保留原脚本", tname, i, n2)
            except Exception as e:
                log.warning("[%s]#%d 扩写失败，保留原脚本: %s", tname, i, e)


def load_mode_lib(card: dict) -> str:
    """读取客户爆款模式库（distill 产出），渲染成 prompt 参考段。"""
    lib = read_json(DATA / str(card["客户"]).strip() / "distill" / "模式库.json") or {}
    modes = lib.get("模式库") or []
    if not modes:
        return ""
    lines = ["【爆款模式库参考】（行业风格灵感，仅用于句式与语气，结构仍以对标视频为准）"]
    for m in modes[:6]:
        lines.append(f"- 模式「{m.get('模式名')}」：{m.get('特征')}")
        lines.append(f"  句式骨架：{m.get('句式骨架')}")
        lines.append(f"  语气：{m.get('语气要点')}｜镜头：{m.get('镜头场景')}")
    return "\n".join(lines)


def _overlap(a: str, b: str) -> int:
    return len(set(str(a)) & set(str(b)))


def build_ref_pool(track: dict, analysis: list, by_id: dict, tdir: Path, n_per: int) -> list:
    """构建赛道参考池：有逐字稿的视频，按（赛道倾向匹配度+代表视频加权+点赞）排序。

    每条脚本绑定池中一条不同的参考视频，保证同赛道脚本各自跟住一条对标。
    """
    tname = track.get("名称", "")
    reps = {str(v) for v in (track.get("代表视频") or [])}
    scored = []
    for a in analysis:
        vid = a.get("video_id")
        if not vid or vid not in by_id:
            continue
        if by_id[vid].get("对照"):
            continue  # 对照组不作改编参考
        tf = tdir / f"{vid}.txt"
        if not tf.exists() or len(tf.read_text(encoding="utf-8").strip()) < 100:
            continue
        tend = str(a.get("赛道倾向", ""))
        s = _overlap(tname, tend) + (3 if vid in reps else 0)
        scored.append((s, -by_id[vid].get("digg_count", 0), vid))
    scored.sort(reverse=True)
    pool = [vid for s, _, vid in scored if s > 0]
    for s, _, vid in scored:
        if len(pool) >= n_per:
            break
        if vid not in pool:
            pool.append(vid)
    return pool[:n_per]


def gen_prompt(track: dict, card: dict, ref_vid: str, by_id: dict,
               analysis_map: dict, transcripts: dict, script_no: int,
               fewshot: str, mode_lib: str = "", used_topics: list = None) -> str:
    """1:1 内容跟随改编：话题/论点/论证方式跟随对标视频，只替换客户特有信息。"""
    lo, hi = parse_duration(card)
    v = by_id.get(ref_vid, {})
    a = analysis_map.get(ref_vid, {})
    transcript = transcripts.get(ref_vid, "")
    used_block = "\n".join(f"- {t}" for t in (used_topics or [])) or "（尚无，本条可自由选题）"
    hook = a.get("钩子设计") or {}
    rhythm = a.get("叙事结构与节奏") or {}
    emotion = a.get("情绪共鸣点") or []
    cta_design = a.get("结尾互动引导") or {}
    cta_lines = "\n".join(f'- "{c}"' for c in _cta_lib(card))
    tags = card.get("客户标签") or []
    persona = card.get("人设档案") or {}
    persona_kuotou = persona.get("口头禅") or []
    persona_block = ""
    if persona:
        persona_block = (f"语气：{persona.get('语气', '')}｜表达习惯：{persona.get('表达习惯', '')}"
                         f"｜口头禅：{'、'.join(persona_kuotou)}")
    scenes = (card.get("客户画像") or {}).get("可拍摄场景") or []
    scene_list = "、".join(str(s) for s in scenes)
    return f"""你是抖音短视频编导。任务：把下面这一条【对标视频】**内容跟随式改编**成一条 {lo}~{hi} 秒的【客户】脚本（本赛道第 {script_no} 条）。

【客户业务】{card.get('业务简介', '').strip()}
【客户卖点】{'、'.join(card.get('卖点') or [])}
【客户人设】{card.get('人设', '')}
【客户标签】{'、'.join(tags) if tags else '（无）'}
【人设档案】{persona_block or '（无）'}
【目标客户】{card.get('目标客户', '')}
【排除规则】{card.get('排除规则', '')}

【赛道】名称:{track.get('名称')}
定位:{track.get('定位')}
常见结构:{track.get('常见结构')}
可拍场景:{track.get('可拍场景')}

【本赛道已用选题（本条必须讲不同侧面，避免同赛道脚本重复）】
{used_block}

【对标视频】（唯一改编对象）
标题：{v.get('desc', '')}
账号：{v.get('author', '')}｜点赞：{v.get('digg_count', 0)}
完整逐字稿：
{transcript}

【对标视频拆解】
钩子设计：类型={hook.get('类型', '')}｜原话={hook.get('前3秒原话', '')}｜抓人机制={hook.get('抓人机制', '')}｜强度={hook.get('强度', '')}
叙事结构与节奏：{json.dumps(rhythm, ensure_ascii=False)}
情绪共鸣点：{json.dumps(emotion, ensure_ascii=False)}
结尾互动引导：{json.dumps(cta_design, ensure_ascii=False)}
爆点归因：{a.get('爆点归因')}
可复用模板：{a.get('可复用模板')}

【改编规则：内容跟随 + 差异化创新】
1. 话题跟随：参考视频讲什么话题，你的脚本就讲什么话题，不得换话题。只替换品牌名、地区、报价方式、联系方式等客户特有信息
2. 论点跟随：核心论点、论证顺序、例子逐条对应保留，换成客户语境下的同类表达
3. 钩子跟随（七型）：前 3 秒按参考视频的钩子类型重写，类型从「反常识/痛点/悬念/冲突/信息差/利益承诺/共鸣」中选，参考视频用什么型就用什么型，禁止"大家好我是…"平铺开场。句式参考：价值型"花了XX元解决了困扰3年的问题"、共鸣型"是不是每次XX都受不了"、悬念型"我发现一个XX行业不告诉你的秘密"
4. 语气跟随 + 个人风格：保留参考视频语气特征，同时严格贴合【客户标签】与【人设档案】——口头禅自然带入口播、用客户的表达习惯（先结论后解释/短句/多用"你"对话），让脚本"像客户本人说出来的话"，而不是通用的营销腔
5. 段落跟随 + 秒级基准：镜头数量与段落对应；参考结构混乱时用基准——0-3s 钩子 → 4-6s 痛点放大 → 7-12s 方案引入 → 13-20s 演示/成果 → 21-26s 数据/前后对比 → 27-35s 价值一句话+互动引导。目标完播率>40%，信息密度前置、无冗余铺垫
6. 红线：允许句子骨架相似但不得连续 20 字逐字照抄；数字/尺寸/价格/材料承诺留白（"以实测为准/私信发资料"）；产品植入放方案段；禁用绝对化用语（"最好/第一/100%有效/全网最低"）
7. 时长 {lo}~{hi} 秒：文案总量 {lo*4}~{hi*4} 字，5~8 个镜头

【拍摄硬规则】（写进分镜表）
- 竖屏 9:16；真人出镜优先（完播率比纯产品画面高 30%+）
- 字幕必配（大量用户静音刷抖音）
- 单镜头 ≤ 15s；景别与内容匹配：特写=情绪/细节，中景=讲解，全景=环境交代

【CTA 选项库】（结尾从中选一种，结合客户卖点）
{cta_lines}

【客户已验证脚本风格示例】
{fewshot}

{mode_lib}

只输出一个 JSON 对象：{{"发布文案":"带3-5个话题标签的发布文案","参考视频":"{ref_vid}","参考主题":"一句话复述参考视频讲什么","改编说明":"必须输出：跟了对标的什么话题/结构，换成了客户的什么（品牌/地区/案例），差异化体现在哪（一句）","镜头":[{{"时间":"0-5s","景别":"远景|全景|中景|近景|特写","场景":"从可拍摄场景[{scene_list}]里选一个","画面":"一句话点明拍什么（≤15字，不写详细运镜）","文案":"口播说什么","拍摄提示":"场景+关键道具，一句话（≤15字）"}}],"字幕建议":"全程字幕+哪些关键词做花字","音效建议":"BGM情绪"}}"""


def ensure_shot_times(scripts_by_track: dict, card: dict):
    """兜底：镜头缺少时间字段时，按文案字数比例自动分配时间段。"""
    lo, hi = parse_duration(card)
    span = hi - lo
    for scripts in scripts_by_track.values():
        for s in scripts:
            shots = s.get("镜头", [])
            if not shots or all(sh.get("时间") for sh in shots):
                continue
            total = sum(len(sh.get("文案", "")) for sh in shots) or 1
            t = lo
            for sh in shots:
                dur = max(3.0, span * len(sh.get("文案", "")) / total)
                sh["时间"] = f"{t:.0f}-{t + dur:.0f}s"
                t += dur


# ---------- 专家方法论注入：合规扫描（绝对化用语禁用，广告法） ----------
_COMPLIANCE_BAN = ("最好", "第一", "100%", "百分百", "全网最低", "绝对", "零差评",
                   "最便宜", "第一名", "顶级", "第一品牌")


def compliance_scan(scripts: list) -> list:
    """合规扫描：绝对化用语禁用。返回 [(脚本序号, 违规词, 片段)]。"""
    warns = []
    for si, s in enumerate(scripts, 1):
        text = s.get("发布文案", "") + "".join(sh.get("文案", "") for sh in s.get("镜头", []))
        for w in _COMPLIANCE_BAN:
            if w in text:
                warns.append((si, w, text[:30]))
                break
    return warns


# ---------- 卖点覆盖矩阵（纯代码统计） ----------
def _sp_keywords(sell: str) -> list:
    return [p.strip() for p in re.split(r"[/／、，,·]", str(sell)) if p.strip()]


def sellpoint_coverage(scripts_by_track: dict, card: dict) -> dict:
    """卖点×脚本覆盖矩阵 + 各卖点覆盖数（零覆盖卖点告警）。"""
    sells = card.get("卖点") or []
    matrix = {}
    for tname, scripts in scripts_by_track.items():
        for i, s in enumerate(scripts, 1):
            text = s.get("发布文案", "") + "".join(sh.get("文案", "") for sh in s.get("镜头", []))
            hit = [sp for sp in sells if any(k and k in text for k in _sp_keywords(sp))]
            matrix[f"{tname}#{i}"] = hit
    coverage = {sp: sum(1 for hits in matrix.values() if sp in hits) for sp in sells}
    return {"matrix": matrix, "coverage": coverage, "卖点数": len(sells)}


# ---------- 发布文案三候选（LLM） ----------
def gen_captions(script: dict, card: dict, api_key: str) -> list:
    """为脚本生成 3 个发布文案候选（悬念式/数据式/痛点式）。"""
    shots = script.get("镜头", [])
    s_hook = shots[0].get("文案", "") if shots else ""
    body = "".join(sh.get("文案", "") for sh in shots)[:300]
    prompt = f"""你是短视频标题专家。下面是客户一条脚本的口播内容，请写 3 个不同风格的发布标题候选。
只输出 JSON：{{"标题":[{{"风格":"悬念式","文案":"..."}},{{"风格":"数据式","文案":"..."}},{{"风格":"痛点式","文案":"..."}}]}}
要求：每个 15-30 字，含钩子，口语化，可带 1-2 个话题标签；不编造数字，禁用"最好/第一/100%"。

脚本首句：{s_hook}
脚本正文（节选）：{body}
客户业务：{card.get('业务简介', '').strip()}"""
    try:
        r = call_deepseek([{"role": "user", "content": prompt}], api_key, temperature=0.8)
        return r.get("标题", [])
    except Exception as e:
        log.warning("发布文案三候选生成失败: %s", e)
        return []


# ---------- LLM Judge：四维评分 + 镜头级定位 ----------
def judge_scripts(scripts_by_track: dict, card: dict, analysis_map: dict, api_key: str) -> dict:
    """四维评分（并发）+ 镜头级问题定位。"""
    out = {}
    items = [(tname, i, s) for tname, scripts in scripts_by_track.items()
             for i, s in enumerate(scripts, 1)]

    def _judge_one(item):
        tname, i, s = item
        ref = analysis_map.get(str(s.get("参考视频", "")), {})
        hook_type = (ref.get("钩子设计") or {}).get("类型", "")
        prompt = f"""你是短视频质量评审。下面是一条改编脚本，请按四维评分（每维 1-10 + 一句依据 + 问题镜头定位）。
只输出 JSON：{{"对标匹配度":1到10,"对标依据":"一句话","原创度":1到10,"原创依据":"一句话","可执行性":1到10,"可执行依据":"一句话","红线合规":1到10,"红线依据":"一句话","问题镜头":[{{"镜头序号":数字,"问题":"钩子弱/台词拖沓/时长失衡/景别不当","建议":"一句话"}}]}}
评分依据：
- 对标匹配度：话题/论点顺序/钩子类型是否跟随对标（对标钩子类型：{hook_type}）
- 原创度：表达是否差异化（换客户语境、本土化），而非逐字照抄
- 可执行性：分镜表能否直接开拍（景别/时长/道具/表演齐备）
- 红线合规：时长达标、无 20 字照抄、无绝对化用语、无事实编造、CTA 在位、钩子非平铺

脚本：{json.dumps(s, ensure_ascii=False)}"""
        try:
            r = call_deepseek([{"role": "user", "content": prompt}], api_key, temperature=0.2)
            r["_四维均分"] = round(
                sum(float(r.get(k, 0) or 0) for k in ("对标匹配度", "原创度", "可执行性", "红线合规")) / 4, 1)
            return tname, i, r
        except Exception as e:
            log.warning("[%s]#%d judge 失败: %s", tname, i, e)
            return tname, i, None

    workers = max(1, min(4, len(items)))
    with cf.ThreadPoolExecutor(max_workers=workers) as ex:
        futs = [ex.submit(_judge_one, it) for it in items]
        for fut in cf.as_completed(futs):
            tname, i, r = fut.result()
            if r:
                scripts_by_track[tname][i - 1]["评审"] = r
                out[f"{tname}#{i}"] = r
    return out


def _similarity(a: str, b: str) -> float:
    """字符串相似度 0-1：rapidfuzz 优先，降级 difflib。"""
    if _rfuzz is not None:
        return _rfuzz.ratio(a, b) / 100.0
    return difflib.SequenceMatcher(None, a, b).ratio()


def check_similarity(scripts: list, transcripts: str, threshold: float = 0.5):
    """查重闸：先查整句包含（快），再全窗口模糊比对（rapidfuzz，不抽样）。"""
    warnings = []
    for si, s in enumerate(scripts):
        for line in s.get("镜头", []):
            text = line.get("文案", "")
            if len(text) < 10:
                continue
            if text in transcripts:  # 整句照搬，直接警告
                warnings.append((si, text[:30]))
                continue
            # 全窗口滑窗（rapidfuzz 快，无需抽样）
            n = len(text)
            step = max(1, n // 2)
            for i in range(0, max(1, len(transcripts) - n + 1), step):
                sub = transcripts[i:i + n]
                if _similarity(text, sub) > threshold:
                    warnings.append((si, text[:30]))
                    break
    return warnings


def structure_similarity(a: dict, b: dict) -> float:
    """结构同质化：各镜头时长分布的相似度（节奏雷同度）。"""
    shots_a = a.get("镜头", [])
    shots_b = b.get("镜头", [])
    if not shots_a or not shots_b:
        return 0.0

    def spans(shots):
        out = []
        for sh in shots:
            m = re.findall(r"\d+", str(sh.get("时间", "")))
            out.append(int(m[1]) - int(m[0]) if len(m) >= 2 else 0)
        return out

    sa, sb = spans(shots_a), spans(shots_b)
    return _similarity("-".join(map(str, sa)), "-".join(map(str, sb)))


def check_structure_dupes(scripts_by_track: dict, threshold: float = 0.95) -> list:
    """同池脚本两两比对结构相似度，>阈值提示"换叙事顺序"。"""
    warns = []
    all_s = [(f"{tname}#{i}", s) for tname, lst in scripts_by_track.items() for i, s in enumerate(lst, 1)]
    for x in range(len(all_s)):
        for y in range(x + 1, len(all_s)):
            name_a, s_a = all_s[x]
            name_b, s_b = all_s[y]
            sim = structure_similarity(s_a, s_b)
            if sim > threshold:
                warns.append((name_a, name_b, round(sim, 2)))
    return warns


def sheet_name(name: str) -> str:
    name = re.sub(r'[\\/*?:\[\]]', "", name).strip()
    return (name or "赛道")[:31]


def make_shooting_sheet(wb, scripts_by_track: dict, card: dict):
    """拍摄执行 sheet：按需求卡可拍摄场景聚合镜头数 + 道具/口播提示清单。"""
    ws = wb.create_sheet("拍摄执行")
    scenes = (card.get("客户画像") or {}).get("可拍摄场景") or []
    all_shots = []
    for tname, scripts in scripts_by_track.items():
        for si, s in enumerate(scripts, 1):
            for shot in s.get("镜头", []):
                all_shots.append({"脚本": f"{tname}#{si}", **shot})

    # 按场景字段归类（镜头显式输出"场景"，准确）
    scene_map = {sc: [] for sc in scenes}
    other = []
    for sh in all_shots:
        sc = sh.get("场景", "")
        matched = sc if sc in scene_map else None
        (scene_map[matched] if matched else other).append(sh)

    ws.cell(1, 1, "拍摄执行包（按场景聚合，同场景镜头可集中一天拍完）").font = _FONT_BOLD
    for ci, h in enumerate(["场景", "镜头数", "涉及脚本"], 1):
        ws.cell(2, ci, h).font = _FONT_BOLD
    r = 3
    for sc, shots in scene_map.items():
        if not shots:
            continue
        ws.cell(r, 1, sc)
        ws.cell(r, 2, len(shots))
        ws.cell(r, 3, "、".join(sorted({sh["脚本"] for sh in shots})))
        r += 1
    if other:
        ws.cell(r, 1, "其他/未归类")
        ws.cell(r, 2, len(other))
        ws.cell(r, 3, "、".join(sorted({sh["脚本"] for sh in other})))
        r += 1

    # 道具/口播提示清单（去重）
    r += 1
    ws.cell(r, 1, "道具与口播提示（跨脚本去重）").font = _FONT_BOLD
    r += 1
    tips = sorted({str(sh.get("拍摄提示", "")).strip()
                   for sh in all_shots if str(sh.get("拍摄提示", "")).strip()})
    for i, t in enumerate(tips, 1):
        ws.cell(r, 1, i)
        ws.cell(r, 2, t)
        r += 1
    for col, w in {"A": 30, "B": 12, "C": 46}.items():
        ws.column_dimensions[col].width = w


# ---------- xlsx 输出：一赛道一 sheet，每条脚本独立对标链接 ----------
import math

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_FONT = Font(name="宋体", size=11)
_FONT_BOLD = Font(name="宋体", size=11, bold=True)
_FONT_LINK = Font(name="宋体", size=9, color="0563C1", underline="single")
_CENTER = Alignment(horizontal="center", vertical="center")
_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_FILL_HEADER = PatternFill("solid", start_color="D9E1F2")   # 表头浅蓝
_FILL_SCRIPT = PatternFill("solid", start_color="FFF2CC")   # 脚本行浅黄


def _disp_len(text: str) -> float:
    """显示宽度估算：中文按 2 个单位，ASCII 按 1。"""
    return sum(2 if ord(ch) > 127 else 1 for ch in text)


def _row_height(text: str, width: float, line_h: float = 16.0, min_h: float = 24.0) -> float:
    """按文字量与列宽精确估算行高，保证换行后全部可见。"""
    lines = max(1, math.ceil(_disp_len(text or "") / max(1.0, width * 0.95)))
    return max(min_h, lines * line_h + 5)


def write_xlsx(out: Path, tracks: list, scripts_by_track: dict, by_id: dict, card: dict = None):
    wb = Workbook()
    wb.remove(wb.active)
    for t in tracks:
        tname = t.get("名称", "赛道")
        ws = wb.create_sheet(sheet_name(tname))
        widths = {"A": 34, "B": 44, "C": 22, "D": 54}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        def style(cell, wrap=True, font=_FONT, fill=None, align=None):
            cell.font = font
            cell.alignment = align or (_CENTER_WRAP if wrap else _CENTER)
            cell.border = _BORDER
            if fill:
                cell.fill = fill

        # 表头（冻结 + 底色加粗）
        for ci, h in enumerate(["对标链接", "发布文案", "画面", "文案"], 1):
            c = ws.cell(row=1, column=ci, value=h)
            style(c, wrap=False, font=_FONT_BOLD, fill=_FILL_HEADER)
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        r = 2
        scripts = scripts_by_track.get(tname, [])
        for si, s in enumerate(scripts, 1):
            # ---- 脚本头行：对标链接 + 带序号的发布文案（浅黄底纹） ----
            vid = str(s.get("参考视频", "")).strip()
            ref = by_id.get(vid, {}).get("url", "")
            c = ws.cell(row=r, column=1, value=ref)
            style(c, wrap=False, font=_FONT_LINK, fill=_FILL_SCRIPT)
            if ref:
                c.hyperlink = ref
            ref_title = by_id.get(vid, {}).get("desc", "")[:20]
            origin = f"对标：{ref_title or vid}" if vid else "原创"
            note = s.get("改编说明", "")
            if not note:
                note = "改编自对标视频，话题/论点跟随，换客户品牌地区" if vid else "原创脚本，思路见发布文案"
            caption = f"脚本{si}｜{origin}｜{note}｜{s.get('发布文案', '')}"
            c2 = ws.cell(row=r, column=2, value=caption)
            style(c2, font=_FONT_BOLD, fill=_FILL_SCRIPT, align=_LEFT_WRAP)
            for ci in (3, 4):
                cc = ws.cell(row=r, column=ci, value="")
                style(cc, fill=_FILL_SCRIPT)
            ws.row_dimensions[r].height = max(24.0, _row_height(caption, widths["B"]))
            r += 1

            # ---- 镜头行 ----
            for shot in s.get("镜头", []):
                pic = shot.get("画面", "")
                text = shot.get("文案", "")
                ts = shot.get("时间", "")
                sc = shot.get("场景", "")
                jing = shot.get("景别", "")
                tags = " ".join(f"[{p}]" for p in (ts, sc, jing) if p)
                pic_show = f"{tags} {pic}".strip() if (tags or pic) else ""
                for ci in (1, 2):
                    cc = ws.cell(row=r, column=ci, value="")
                    style(cc)
                c3 = ws.cell(row=r, column=3, value=pic_show)
                style(c3)
                c4 = ws.cell(row=r, column=4, value=text)
                style(c4, align=_LEFT_WRAP)
                ws.row_dimensions[r].height = max(
                    _row_height(pic_show, widths["C"]), _row_height(text, widths["D"]))
                r += 1

            # ---- 脚本尾行：字幕/音效建议 ----
            sub = s.get("字幕建议", "")
            snd = s.get("音效建议", "")
            if sub or snd:
                for ci in (1, 3, 4):
                    cc = ws.cell(row=r, column=ci, value="")
                    style(cc)
                c2 = ws.cell(row=r, column=2, value=f"字幕：{sub or '无'}｜音效：{snd or '无'}")
                style(c2, font=_FONT, align=_LEFT_WRAP)
                ws.row_dimensions[r].height = _row_height(f"字幕：{sub or '无'}｜音效：{snd or '无'}", widths["B"] + widths["C"])
                r += 1

            # ---- 脚本间隔行（空白分隔，无边框） ----
            if si < len(scripts):
                ws.row_dimensions[r].height = 14
                r += 1

    # 拍摄执行包（按场景聚合 + 道具清单）
    if card:
        make_shooting_sheet(wb, scripts_by_track, card)

    try:
        wb.save(out)
        log.info("脚本池已保存: %s（一赛道一 sheet，脚本间隔+自适应行高）", out)
    except PermissionError:
        from datetime import datetime
        alt = out.with_name(out.stem + "_" + datetime.now().strftime("%H%M%S") + ".xlsx")
        wb.save(alt)
        log.warning("脚本池.xlsx 被占用（可能正被 Excel 打开），已保存到: %s", alt)


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("card")
    ap.add_argument("--resume", action="store_true", help="兼容 run.py 调用")
    ap.add_argument("--fresh", action="store_true", help="新建运行目录（不复用上次）")
    ap.add_argument("--force", action="store_true", help="重新生成全部脚本（调 LLM）")
    ap.add_argument("--xlsx-only", action="store_true", help="只从 scripts.json 重写 xlsx（不调 LLM）")
    ap.add_argument("--fix", action="store_true", help="只做时长兜底扩写+重写输出（调 LLM 扩写短脚本）")
    ap.add_argument("--wave", choices=["test", "bulk"], default="bulk",
                    help="test=测试波(每赛道1条，一周内验证方向)；bulk=量产(按需求卡数量)")
    args = ap.parse_args()

    card = load_card(args.card)
    d = run_dir(card, resume=not args.fresh)
    setup_log(d / "run.log")

    out = d / "脚本池.xlsx"

    if args.xlsx_only:
        tracks = (read_json(d / "tracks.json") or {}).get("赛道", [])
        cand = read_json(d / "candidates.json") or {}
        by_id = {v["aweme_id"]: v for v in cand.get("视频", [])}
        scripts_by_track = read_json(d / "scripts.json") or {}
        if not tracks or not scripts_by_track:
            sys.exit("tracks.json / scripts.json 不存在，无法只重写 xlsx")
        write_xlsx(out, tracks, scripts_by_track, by_id, card)
        return

    api_key = get_api_key()
    if not api_key:
        sys.exit("缺少 LLM_API_KEY / DEEPSEEK_API_KEY 环境变量")

    tracks = (read_json(d / "tracks.json") or {}).get("赛道", [])
    analysis = (read_json(d / "analysis.json") or {}).get("视频分析", [])
    cand = read_json(d / "selected_candidates.json") or read_json(d / "candidates.json") or {}
    # 聚焦模式：需求卡指定主攻赛道时，只生成该赛道
    focus = str(card.get("生成设置", {}).get("主攻赛道", "") or "").strip()
    if focus:
        tracks = [t for t in tracks if t.get("名称") == focus]
        if not tracks:
            sys.exit(f"主攻赛道[{focus}] 不在聚类结果里，请检查 tracks.json")
        log.info("聚焦模式：只生成主攻赛道 [%s]", focus)
    if not tracks:
        sys.exit("tracks.json 不存在，请先跑 cluster")
    by_id = {v["aweme_id"]: v for v in cand.get("视频", [])}

    if args.fix:
        track_names = {t.get("名称", "") for t in tracks}
        scripts_by_track = {
            k: v for k, v in (read_json(d / "scripts.json") or {}).items()
            if k in track_names
        }
        if not scripts_by_track:
            sys.exit("scripts.json 不存在，无法修复")
        analysis_map = {a.get("video_id"): a for a in analysis}
        fix_lengths(scripts_by_track, card, api_key)
        fix_hooks(scripts_by_track, analysis_map, card, api_key)
        ensure_shot_times(scripts_by_track, card)
        write_json(d / "scripts.json", scripts_by_track)
        write_xlsx(out, tracks, scripts_by_track, by_id, card)
        return

    if out.exists() and not args.force:
        log.info("脚本池.xlsx 已存在，跳过生成（--force 重跑）")
        return

    fewshot_path = card.get("生成设置", {}).get("fewshot_脚本xlsx", "")
    examples = parse_fewshot_xlsx(fewshot_path)
    fewshot_text = json.dumps(examples[:2], ensure_ascii=False, indent=2) if examples else "（无）"

    n_per = int(card.get("生成设置", {}).get("每赛道脚本数", 5))
    if args.wave == "test":
        n_per = 1
        log.info("测试波模式：每赛道仅生成 1 条，一周内回收数据定向后再跑 --wave bulk 量产")
    tdir = d / "transcripts"
    # 断点续跑：已生成过的赛道跳过 LLM 调用；不在当前 tracks 里的旧赛道丢弃
    track_names = {t.get("名称", "") for t in tracks}
    old_all = read_json(d / "scripts.json") or {}
    scripts_by_track = {
        k: v for k, v in old_all.items()
        if k in track_names
    }
    if args.force:
        scripts_by_track = {}  # 强制重新生成全部脚本

    def _gen_track(t):
        """并发生成一个赛道的全部脚本（赛道内串行保持选题去重）。"""
        tname = t.get("名称", "赛道")
        if tname in scripts_by_track:
            return tname, None, "已有脚本跳过"
        pool = build_ref_pool(t, analysis, by_id, tdir, n_per)
        if not pool:
            return tname, None, "没有可用参考视频"
        transcripts = {}
        for vid in pool:
            tf = tdir / f"{vid}.txt"
            transcripts[vid] = tf.read_text(encoding="utf-8").strip() if tf.exists() else ""
        analysis_map = {a.get("video_id"): a for a in analysis}
        mode_lib = load_mode_lib(card)
        track_scripts = []
        used_topics = []
        for idx in range(n_per):
            ref_vid = pool[idx % len(pool)]
            try:
                prompt = gen_prompt(t, card, ref_vid, by_id, analysis_map,
                                    transcripts, idx + 1, fewshot_text, mode_lib,
                                    used_topics)
                result = call_deepseek([{"role": "user", "content": prompt}],
                                       api_key, temperature=0.8)
                # 兼容两种返回：{"脚本":[...]} 或 直接是单条脚本对象
                script = None
                if isinstance(result.get("脚本"), list) and result.get("脚本"):
                    script = result["脚本"][0]
                elif result.get("镜头"):
                    script = result
                if not script or not script.get("镜头"):
                    raise ValueError("LLM 未返回有效脚本结构")
                script["参考视频"] = ref_vid
                script["改编角度"] = script.get("参考主题", "")[:30] or f"改编自{ref_vid}"
                track_scripts.append(script)
                used_topics.append(script["改编角度"])
                log.info("赛道[%s] 脚本%d 生成成功（参考 %s）", tname, idx + 1, ref_vid)
            except Exception as e:
                log.error("赛道[%s] 脚本%d 生成失败: %s", tname, idx + 1, e)
        if len(track_scripts) < n_per and tname in old_all:
            log.error("赛道[%s] 生成不完整(%d/%d)，保留旧脚本", tname, len(track_scripts), n_per)
            track_scripts = old_all[tname]
        return tname, track_scripts, ""

    # 赛道级并发生成（3 赛道并行，每赛道内串行保持选题去重）
    todo_tracks = [t for t in tracks if t.get("名称") not in scripts_by_track]
    workers = max(1, min(3, len(todo_tracks)))
    with cf.ThreadPoolExecutor(max_workers=workers) as ex:
        futs = [ex.submit(_gen_track, t) for t in todo_tracks]
        for fut in cf.as_completed(futs):
            tname, track_scripts, err = fut.result()
            if err:
                log.error("赛道[%s] %s", tname, err)
                continue
            if track_scripts:
                scripts_by_track[tname] = track_scripts
                log.info("赛道[%s] 共 %d 条脚本", tname, len(track_scripts))

    if not scripts_by_track:
        log.error("全部赛道生成失败，未覆盖旧结果（scripts.json 与 xlsx 保持不变）")
        sys.exit(1)
    analysis_map = {a.get("video_id"): a for a in analysis}
    # 时长兜底：过短脚本自动扩写/超长自动裁剪
    fix_lengths(scripts_by_track, card, api_key)
    # 钩子兜底：平铺开场按参考钩子类型重写
    fix_hooks(scripts_by_track, analysis_map, card, api_key)
    # 镜头时间戳兜底：缺时间字段的按字数比例自动分配
    ensure_shot_times(scripts_by_track, card)

    # ---- 专家注入：合规扫描（绝对化用语，广告法） ----
    all_scripts = [s for lst in scripts_by_track.values() for s in lst]
    comp_warns = compliance_scan(all_scripts)
    if comp_warns:
        log.warning("合规扫描：%d 处绝对化用语需修改: %s", len(comp_warns),
                    [(w, t) for _, w, t in comp_warns[:5]])

    # ---- 卖点覆盖矩阵 ----
    cov = sellpoint_coverage(scripts_by_track, card)
    write_json(d / "卖点覆盖.json", cov)
    zero = [sp for sp, n in cov["coverage"].items() if n == 0]
    if zero:
        log.warning("卖点覆盖：%d 个卖点零覆盖（建议下轮定向注入）: %s", len(zero), zero)
    else:
        log.info("卖点覆盖矩阵已生成，%d 个卖点全部被覆盖", cov["卖点数"])

    # ---- 发布文案三候选（并发） ----
    _cap_items = [(tname, si, s) for tname, lst in scripts_by_track.items()
                  for si, s in enumerate(lst)]

    def _cap_one(item):
        tname, si, s = item
        return tname, si, gen_captions(s, card, api_key)

    _cap_workers = max(1, min(4, len(_cap_items)))
    with cf.ThreadPoolExecutor(max_workers=_cap_workers) as ex:
        _futs = [ex.submit(_cap_one, it) for it in _cap_items]
        for _fut in cf.as_completed(_futs):
            tname, si, caps = _fut.result()
            if caps:
                scripts_by_track[tname][si]["标题候选"] = caps

    # ---- LLM Judge：四维评分 + 镜头级定位 ----
    judge_scripts(scripts_by_track, card, analysis_map, api_key)

    # 元信息兜底：LLM 未返回的脚本级字段补默认值
    for lst in scripts_by_track.values():
        for s in lst:
            s.setdefault("字幕建议", "全程字幕+关键词花字")
            s.setdefault("音效建议", "轻快BGM")

    write_json(d / "scripts.json", scripts_by_track)
    write_xlsx(out, tracks, scripts_by_track, by_id, card)

    # 查重闸汇总
    all_text = "".join(
        (tdir / f"{v['aweme_id']}.txt").read_text(encoding="utf-8")
        for v in by_id.values() if (tdir / f"{v['aweme_id']}.txt").exists()
    )
    warns = check_similarity(
        [s for lst in scripts_by_track.values() for s in lst], all_text
    )
    if warns:
        log.warning("查重提示：%d 句与参考逐字稿相似度过高，建议人工改写: %s",
                    len(warns), warns[:5])

    # 结构同质化检查（防"换词不换骨"）
    dupes = check_structure_dupes(scripts_by_track)
    if dupes:
        log.warning("结构同质化：%d 对脚本节奏雷同（建议换叙事顺序）: %s",
                    len(dupes), dupes[:5])


if __name__ == "__main__":
    main()

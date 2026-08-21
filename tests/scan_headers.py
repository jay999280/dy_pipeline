# -*- coding: utf-8 -*-
"""扫描其他客户文件里的表头关键字位置。"""
from openpyxl import load_workbook

for name, path in {
    "凯恩怡家": r"凯恩怡家脚本0506.xlsx",
    "钢栈桥": r"钢栈桥脚本7.29.xlsx",
}.items():
    wb = load_workbook(path, read_only=True, data_only=True)
    print(f"===== {name}")
    for ws in wb.worksheets:
        hits = []
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            for ci, c in enumerate(row):
                if c is not None and str(c).strip() in ("画面", "文案", "旁白", "镜头", "发布文案", "内容", "口播"):
                    hits.append((ri, ci, str(c).strip()))
        print(f"  sheet[{ws.title}] max_row={ws.max_row} 关键字位置: {hits[:10]}")
    wb.close()

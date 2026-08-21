# -*- coding: utf-8 -*-
"""验证 xlsx 显示优化：行高容量 vs 文字量、间隔行、表头样式。"""
import math
from openpyxl import load_workbook

path = r"dy_pipeline/data\示例客户\run_20260815_130550\脚本池.xlsx"
wb = load_workbook(path)

def disp_len(s):
    return sum(2 if ord(ch) > 127 else 1 for ch in str(s))

overflow = []
for name in wb.sheetnames:
    ws = wb[name]
    widths = {"A": 34, "B": 44, "C": 22, "D": 54}
    col_letters = ["A", "B", "C", "D"]
    spacer_rows = []
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if c.value:
                h = ws.row_dimensions[c.row].height or 15
                cap_lines = math.floor(h / 16)  # 每行约 16pt
                w = widths.get(c.column_letter, 20)
                need_lines = math.ceil(disp_len(c.value) / max(1, w * 0.95))
                if need_lines > cap_lines:
                    overflow.append((name, c.coordinate, str(c.value)[:20], need_lines, cap_lines))
        if all(c.value in (None, "") for c in row):
            spacer_rows.append(row[0].row)
    print(f"{name}: 空行(间隔) at rows {spacer_rows} | freeze={ws.freeze_panes} | "
          f"表头填充={ws['A1'].fill.start_color.rgb} 粗体={ws['A1'].font.bold}")

print(f"\n行高不足(显示不全)的单元格: {len(overflow)} 个")
for o in overflow[:10]:
    print("  ", o)
wb.close()
print("检查完成")

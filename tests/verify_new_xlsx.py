# -*- coding: utf-8 -*-
"""对比新生成的 xlsx 与示例模板的结构。"""
from openpyxl import load_workbook

new = r"dy_pipeline/data\示例客户\run_20260815_130550\脚本池_新.xlsx"
wb = load_workbook(new)
print("sheet 数:", len(wb.sheetnames), "|", wb.sheetnames[:6], "...")
ws = wb[wb.sheetnames[0]]
print("=== 第一个 sheet 结构 ===")
print("A1:", ws["A1"].value, "| B1:", str(ws["B1"].value)[:40], "| hyperlink:", bool(ws["B1"].hyperlink))
print("A2:", ws["A2"].value, "| B2(合并):", str(ws["B2"].value)[:45])
print("合并区域:", [str(m) for m in ws.merged_cells.ranges][:8])
print("A3:", ws["A3"].value, "| C3:", ws["C3"].value)
print("字体:", ws["A1"].font.name, ws["A1"].font.sz, "| 边框:", ws["A1"].border.left.style)
print("列宽 A/B/C:", round(ws.column_dimensions["A"].width,1), round(ws.column_dimensions["B"].width,1), round(ws.column_dimensions["C"].width,1))
print("=== 镜头区前 6 行 ===")
for row in ws.iter_rows(min_row=4, max_row=9, max_col=3):
    print("  ", [(c.value if c.value is not None else "")[:20] for c in row])
wb.close()

# -*- coding: utf-8 -*-
"""打印示例脚本 xlsx 前两个 sheet 的真实布局。"""
from openpyxl import load_workbook

path = r"示例脚本.xlsx"
wb = load_workbook(path, read_only=True, data_only=True)
for ws in wb.worksheets[:2]:
    print(f"\n===== sheet: {ws.title} | max_row={ws.max_row} max_col={ws.max_column}")
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [str(c).replace("\n", "\\n")[:40] if c is not None else "" for c in row]
        if any(cells):
            print(f"R{ri}: {cells}")
        if ri >= 25:
            break
wb.close()

# -*- coding: utf-8 -*-
"""打印另外两个客户 xlsx 的布局。"""
from openpyxl import load_workbook

for name, path in {
    "凯恩怡家": r"凯恩怡家脚本0506.xlsx",
    "钢栈桥": r"钢栈桥脚本7.29.xlsx",
}.items():
    wb = load_workbook(path, read_only=True, data_only=True)
    print(f"\n===== {name} | sheets: {wb.sheetnames[:8]}")
    ws = wb.worksheets[0]
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        cells = [str(c).replace("\n", "\\n")[:45] if c is not None else "" for c in row]
        if any(cells):
            print(f"R{ri}: {cells[:6]}")
        if ri >= 20:
            break
    wb.close()

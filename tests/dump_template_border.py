# -*- coding: utf-8 -*-
"""确认模板的边框样式和对标链接位置。"""
from openpyxl import load_workbook

wb = load_workbook(r"示例脚本.xlsx")
ws = wb["服务介绍"]
print("A1 超链接:", ws["A1"].hyperlink, "| B1:", repr(ws["B1"].value), "| B1 hyperlink:", ws["B1"].hyperlink)
for coord in ("A1", "B1", "A2", "B2", "A3", "C3", "B4", "C4", "A5", "B5", "C5"):
    c = ws[coord]
    b = c.border
    print(f"{coord}: left={b.left.style} right={b.right.style} top={b.top.style} bottom={b.bottom.style} "
          f"font={c.font.name},{c.font.sz}")
wb.close()

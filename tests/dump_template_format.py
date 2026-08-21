# -*- coding: utf-8 -*-
"""深挖示例模板 xlsx 的格式细节：合并单元格、列宽、行高、样式。"""
from openpyxl import load_workbook

path = r"示例脚本.xlsx"
wb = load_workbook(path)  # 非 read_only，能读样式
ws = wb["不锈钢介绍"]
print("sheet:", ws.title, "| dims:", ws.dimensions)
print("=== merged cells ===")
for mc in ws.merged_cells.ranges:
    print("  ", str(mc))
print("=== column widths ===")
for k, v in ws.column_dimensions.items():
    print(f"  col {k}: width={v.width}")
print("=== row heights (first 12) ===")
for i in range(1, 13):
    rd = ws.row_dimensions.get(i)
    print(f"  row {i}: height={rd.height if rd else None}")
print("=== cell details (A1:D6) ===")
for row in ws["A1:D6"]:
    for c in row:
        if c.value is not None:
            fill = c.fill.fgColor.rgb if c.fill and c.fill.patternType else None
            print(f"  {c.coordinate}: '{str(c.value)[:24]}' font_bold={c.font.bold} font_size={c.font.sz} "
                  f"fill={fill} wrap={c.alignment.wrap_text} halign={c.alignment.horizontal} valign={c.alignment.vertical}")
wb.close()

# -*- coding: utf-8 -*-
"""脚本池 xlsx 输出测试：合成数据写盘再读回验证结构。"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
from generate import write_xlsx, sheet_name, check_similarity
from openpyxl import load_workbook

out = Path(__file__).parent / "_test_out.xlsx"
tracks = [
    {"名称": "装修避坑/厨房决策", "代表视频": ["111"]},
    {"名称": "材料对比", "代表视频": ["222"]},
]
scripts_by_track = {
    "装修避坑/厨房决策": [
        {"发布文案": "测试文案#标签", "镜头": [{"画面": "门口讲", "文案": "大家好"}, {"画面": "展厅讲", "文案": "结尾CTA"}]},
    ],
    "材料对比": [
        {"发布文案": "第二条", "镜头": [{"画面": "展柜", "文案": "对比讲解"}]},
    ],
}
by_id = {
    "111": {"url": "https://www.douyin.com/video/111"},
    "222": {"url": "https://www.douyin.com/video/222"},
}

write_xlsx(out, tracks, scripts_by_track, by_id)
wb = load_workbook(out, read_only=True)
names = wb.sheetnames
print("sheet 名:", names)
ws = wb[names[0]]
rows = list(ws.iter_rows(values_only=True))
print("表头:", rows[0][:4])
print("脚本行:", rows[1][:4])
print("镜头行:", rows[2][:4])
wb.close()
out.unlink()

# sheet 名清洗
assert sheet_name('装修避坑/厨房决策?') == "装修避坑厨房决策"
# 查重闸
warns = check_similarity(
    [{"镜头": [{"文案": "这是完全照搬的一句原文"}]}],
    "这是完全照搬的一句原文，后面还有别的内容",
)
assert warns, "整句照搬应被查重闸捕获"
assert not check_similarity([{"镜头": [{"文案": "全新的原创句子没有重复"}]}], "别的完全不相关的内容")
print("write_xlsx / sheet_name / 查重闸 测试通过")

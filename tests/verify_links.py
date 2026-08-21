# -*- coding: utf-8 -*-
"""验证：每条脚本的参考视频链接是否正确、是否有重复。"""
import json
from pathlib import Path

from openpyxl import load_workbook

d = Path(r"dy_pipeline/data\示例客户\run_20260815_130550")
sc = json.load(open(d / "scripts.json", encoding="utf-8"))
cand = {v["aweme_id"]: v for v in json.load(open(d / "candidates.json", encoding="utf-8"))["视频"]}

print("=== scripts.json 参考视频 ===")
all_refs = []
for track, scripts in sc.items():
    for i, s in enumerate(scripts, 1):
        vid = s.get("参考视频", "")
        v = cand.get(str(vid), {})
        all_refs.append((track, str(vid)))
        print(f"  [{track}#{i}] {vid} | 有效={bool(v)} | {v.get('desc','')[:22]} (赞{v.get('digg_count',0)})")

print("\n=== xlsx 结构 ===")
wb = load_workbook(d / "脚本池.xlsx")
print("sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    links = [r[0] for r in rows[1:] if r[0]]
    print(f"  {name}: {len(rows)-1} 行 | 脚本行 {len(links)} 条 | 链接去重后 {len(set(links))} 个")
    for lk in links:
        vid = lk.split("/video/")[-1]
        v = cand.get(vid, {})
        print(f"      {vid} | {v.get('desc','')[:26]}")
wb.close()

# 全局检查：链接是否有效、是否跨赛道重复
valid = all(vid in cand for _, vid in all_refs)
print(f"\n全部链接有效: {valid} | 全局唯一链接数: {len(set(v for _, v in all_refs))}/{len(all_refs)}")
